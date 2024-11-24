# Импорты
import numpy as np
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill
from openpyxl.utils.dataframe import dataframe_to_rows
from config import *


# Функция для мэтчинга данных
def match_beeps_date(beep_df, answers_df):

    # 1. Ивзлекаем объекты для сравнения
    beep_df['sent_date'] = beep_df['sent_dt'].dt.date
    beep_df['sent_time'] = beep_df['sent_dt'].dt.time
    answers_df['finished_date'] = answers_df['finished_dt'].dt.date
    answers_df['finished_time'] = answers_df['finished_dt'].dt.time

    # 2. Объединяем данные
    merged_df = beep_df.merge(
        answers_df,
        on=['id', 'survey_name'],
        how='left',
    )

    # 3. Убираем ложные мэтчи
    condition = (
            ((merged_df['finished_date'] == merged_df['sent_date']) & (
                        merged_df['finished_time'] > merged_df['sent_time'])) |
            ((merged_df['finished_date'] == merged_df['sent_date'] + pd.Timedelta(days=1)) & (
                        merged_df['finished_time'] < merged_df['sent_time']))
    )
    merged_df.loc[~condition, "finished_dt"] = pd.NA

    # 4. Проставляем статус завершения
    merged_df['complete'] = np.where(merged_df['finished_dt'].isna(), False, True)

    # 5. Убираем неподходящие опросы
    surveys = [item[0] for item in ANKET_SURVEYS_LIST]
    merged_df = merged_df[merged_df['survey_name'].isin(surveys)]

    # 6. Возвращаем замэтченные данные
    merged_df.to_excel('merged.xlsx')
    return merged_df


# Функция для суммирования данных
def summarize_activity(merged_df):

    # Добавляем колонку с днем в формате "день.месяц"
    merged_df['day_month'] = merged_df['sent_dt'].dt.strftime('%d.%m')

    # Считаем сумму complete == True и общее количество строк по id и day_month
    grouped = merged_df.groupby(['id', 'day_month'])['complete'].agg(
        [('sum_true', 'sum'), ('count', 'count')]).reset_index()

    # Вычисляем долю: сумма True / общее количество строк
    grouped['ratio'] = grouped['sum_true'] / grouped['count']

    # Переходим к широкой форме
    wide_df = grouped.pivot(index='id', columns='day_month', values='ratio').fillna(0)

    # Добавляем колонку average с средним значением по строке
    wide_df['total'] = wide_df.mean(axis=1)

    # Добавляем строку average с средним значением по колонкам
    wide_df.loc['total'] = wide_df.mean(axis=0)

    # Перемещаем колонку total сразу после id
    columns_order = ['total'] + [col for col in wide_df.columns if col not in ['total']]
    wide_df = wide_df[columns_order]

    # Возвращаем таблицу с Summary
    return wide_df


# Функция для форматирования табличек и записи их в XLSX-файл
def export_results(summarized_df, merged_df, file_name="summary.xlsx"):

    # Создаем новую книгу
    wb = Workbook()

    # ========== Первый лист: Compliance Rate (CR) ==========
    ws1 = wb.active
    ws1.title = "Compliance Rate (CR)"

    # Предобработка датафрейма
    summarized_df = summarized_df.reset_index().rename(columns={"index": "id"})

    for r_idx, row in enumerate(dataframe_to_rows(summarized_df, index=False, header=True), start=1):
        for c_idx, value in enumerate(row, start=1):
            cell = ws1.cell(row=r_idx, column=c_idx, value=value)

            # Форматируем первую строку и первую колонку как жирные
            if r_idx == 1 or c_idx == 1:
                cell.font = Font(bold=True)

            # Окрашиваем ячейки на основе значений (от 0 - красный до 1 - зеленый)
            elif isinstance(value, (int, float)) and 0 <= value <= 1:
                green = int(value * 255)
                red = 255 - green
                cell.fill = PatternFill(
                    start_color=f"{red:02X}{green:02X}00",
                    end_color=f"{red:02X}{green:02X}00",
                    fill_type="solid"
                )

    # ========== Второй лист: Prompts ==========
    ws2 = wb.create_sheet(title="Prompts")

    # Убираем лишние колонки
    merged_df.drop(columns=['finished_dt', 'sent_dt', 'day_month'], inplace=True)

    for r_idx, row in enumerate(dataframe_to_rows(merged_df, index=False, header=True), start=1):
        for c_idx, value in enumerate(row, start=1):
            cell = ws2.cell(row=r_idx, column=c_idx, value=value)

            # Форматируем первую строку и первую колонку как жирные
            if r_idx == 1 or c_idx == 1:
                cell.font = Font(bold=True)

            # Проверяем, относится ли текущий столбец к 'complete'
            # Учитываем, что индекс добавляется как первая колонка
            if c_idx - 1 >= 0 and c_idx - 1 < len(merged_df.columns) and merged_df.columns[c_idx - 1] == 'complete':
                if isinstance(value, bool):
                    color = "00FF00" if value else "FF0000"  # Зеленый для True, красный для False
                    cell.fill = PatternFill(
                        start_color=color,
                        end_color=color,
                        fill_type="solid"
                    )

    # Сохраняем книгу в файл
    wb.save("summary.xlsx")
